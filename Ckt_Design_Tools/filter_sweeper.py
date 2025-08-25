#!/usr/bin/env python3
"""
filter Sweeper

"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np
import matplotlib
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib import cm
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt
from math import pi, sqrt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import datetime
from PIL import Image as PILImage

# ---------- Theme / Colors ----------
BG = "#032B52"
FRAME_BG = "#053A6A"
INPUT_YELLOW = "#fff06b"
BTN_BG = "#0A74B7"
BTN_LIGHT = "#F0E68C"
TEXT_FG = "white"
INPUT_FG = "black"
FONT_UI = ("Segoe UI", 10)

matplotlib.rcParams['font.size'] = 10
matplotlib.rcParams['axes.facecolor'] = "#0e2f4f"
matplotlib.rcParams['figure.facecolor'] = "#053A6A"
matplotlib.rcParams['savefig.facecolor'] = "#053A6A"

# ---------- Helpers ----------
def fmt_number(v):
    if v == 0:
        return "0.000"
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "-"
    if abs(v) >= 1e6 or abs(v) <= 1e-3:
        return f"{v:.3e}"
    return f"{v:.3f}"

def human_freq(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "N/A"
    if v >= 1e6:
        return f"{v/1e6:.3f} MHz"
    if v >= 1e3:
        return f"{v/1e3:.3f} kHz"
    return f"{v:.3f} Hz"

# ---------- Math ----------
def fc_RC(R, C):
    return 1.0 / (2 * pi * R * C) if R > 0 and C > 0 else np.nan

def fc_RL(R, L):
    return R / (2 * pi * L) if R > 0 and L > 0 else np.nan

def f0_RLC(L, C):
    return 1.0 / (2 * pi * sqrt(L * C)) if L > 0 and C > 0 else np.nan

def BW_RLC(R, L):
    return R / (2 * pi * L) if L > 0 else np.nan

def Q_RLC(R, L, C):
    bw = BW_RLC(R, L)
    f0 = f0_RLC(L, C)
    return f0 / bw if bw and not np.isnan(bw) and bw != 0 else np.nan

# Transfer functions (mag & phase)
def H_RC_lowpass(R, C, f):
    w = 2*pi*f
    H = 1.0 / np.sqrt(1 + (w*R*C)**2)
    phi = -np.degrees(np.arctan(w*R*C))
    return H, phi

def H_RC_highpass(R, C, f):
    w = 2*pi*f
    H = (w*R*C) / np.sqrt(1 + (w*R*C)**2)
    phi = np.degrees(np.pi/2 - np.arctan(w*R*C))
    return H, phi

def H_RL_lowpass(R, L, f):
    w = 2*pi*f
    H = 1.0 / np.sqrt(1 + (w*L/R)**2)
    phi = -np.degrees(np.arctan(w*L/R))
    return H, phi

def H_RL_highpass(R, L, f):
    w = 2*pi*f
    H = (w*L/R) / np.sqrt(1 + (w*L/R)**2)
    phi = np.degrees(np.pi/2 - np.arctan(w*L/R))
    return H, phi

# ---------- App ----------
class FilterTool3DApp:
    def __init__(self, root):
        self.root = root
        root.title("Filter Selector & Analyzer — Resizable Panel")
        root.configure(bg=BG)
        try:
            root.state('zoomed')
        except Exception:
            pass

        self.main = tk.Frame(root, bg=FRAME_BG, padx=12, pady=12)
        self.main.pack(expand=True, fill="both")

        # build UI pieces
        self.build_component_group()   # top area with R/C/L + result (top-right)
        self.build_freq_group()
        self.build_action_group()
        self.build_plot_tabs()
        self.build_table_and_info_with_pane()  # uses PanedWindow so resizable
        self.build_3d_controls()

        # caches
        self.last_var_vals = None
        self.last_fc = None
        self.last_bw = None
        self.last_q = None
        self.last_gain = None
        self.last_phase = None
        self.last_var_name = None
        self.formula_text_str = ""
        self.export_ready = False

        self.clear_plots()

    # -------- UI: top component group (with result to the right) ----------
    def build_component_group(self):
        frame = tk.LabelFrame(self.main, text="Component Values", bg=FRAME_BG, fg=TEXT_FG, font=FONT_UI, padx=8, pady=8)
        frame.pack(fill="x", pady=(0,8))

        lbl_font = ("Segoe UI", 11)
        entry_w = 16

        # Mode & Filter type
        tk.Label(frame, text="Mode:", bg=FRAME_BG, fg=TEXT_FG, font=lbl_font).grid(row=0, column=0, sticky="e", padx=(4,6))
        self.mode_var = tk.StringVar(value="RC")
        mode_menu = ttk.Combobox(frame, textvariable=self.mode_var, values=["RC","RL","RLC"], width=6, state="readonly")
        mode_menu.grid(row=0, column=1, sticky="w", padx=(0,12))

        tk.Label(frame, text="Filter:", bg=FRAME_BG, fg=TEXT_FG, font=lbl_font).grid(row=0, column=2, sticky="e", padx=(4,6))
        self.filter_type_var = tk.StringVar(value="Low-Pass")
        fmenu = ttk.Combobox(frame, textvariable=self.filter_type_var, values=["Low-Pass","High-Pass"], width=14, state="readonly")
        fmenu.grid(row=0, column=3, sticky="w", padx=(0,12))

        # R
        tk.Label(frame, text="R:", bg=FRAME_BG, fg=TEXT_FG, font=lbl_font).grid(row=1, column=0, sticky="e", padx=(4,6), pady=(8,0))
        self.r_entry = tk.Entry(frame, width=entry_w, bg=INPUT_YELLOW, fg=INPUT_FG, font=("Helvetica",12), insertbackground=INPUT_FG)
        self.r_entry.insert(0, "1.000")
        self.r_entry.grid(row=1, column=1, sticky="w", padx=(0,6), pady=(8,0))
        self.r_unit = tk.StringVar(value="MΩ")
        self.r_unit_menu = tk.OptionMenu(frame, self.r_unit, "Ω","kΩ","MΩ"); self.r_unit_menu.config(bg=INPUT_YELLOW, fg=INPUT_FG, width=6)
        self.r_unit_menu.grid(row=1, column=2, sticky="w", padx=(0,12), pady=(8,0))

        # C
        tk.Label(frame, text="C:", bg=FRAME_BG, fg=TEXT_FG, font=lbl_font).grid(row=2, column=0, sticky="e", padx=(4,6), pady=(8,0))
        self.c_entry = tk.Entry(frame, width=entry_w, bg=INPUT_YELLOW, fg=INPUT_FG, font=("Helvetica",12), insertbackground=INPUT_FG)
        self.c_entry.insert(0, "22.000")
        self.c_entry.grid(row=2, column=1, sticky="w", padx=(0,6), pady=(8,0))
        self.c_unit = tk.StringVar(value="nF")
        self.c_unit_menu = tk.OptionMenu(frame, self.c_unit, "pF","nF","μF"); self.c_unit_menu.config(bg=INPUT_YELLOW, fg=INPUT_FG, width=6)
        self.c_unit_menu.grid(row=2, column=2, sticky="w", padx=(0,12), pady=(8,0))

        # L
        tk.Label(frame, text="L:", bg=FRAME_BG, fg=TEXT_FG, font=lbl_font).grid(row=3, column=0, sticky="e", padx=(4,6), pady=(8,0))
        self.l_entry = tk.Entry(frame, width=entry_w, bg=INPUT_YELLOW, fg=INPUT_FG, font=("Helvetica",12), insertbackground=INPUT_FG)
        self.l_entry.insert(0, "10.000")
        self.l_entry.grid(row=3, column=1, sticky="w", padx=(0,6), pady=(8,0))
        self.l_unit = tk.StringVar(value="mH")
        self.l_unit_menu = tk.OptionMenu(frame, self.l_unit, "H","mH","μH","nH"); self.l_unit_menu.config(bg=INPUT_YELLOW, fg=INPUT_FG, width=6)
        self.l_unit_menu.grid(row=3, column=2, sticky="w", padx=(0,12), pady=(8,0))

        # Make a spot on the right for the Calculate results - top-right by R/C/L
        self.top_result_label = tk.Label(frame, text="", bg=FRAME_BG, fg="cyan", font=("Segoe UI", 12, "bold"), justify="left")
        # Place it spanning the 4 rows, to the right
        self.top_result_label.grid(row=0, column=4, rowspan=4, sticky="nsew", padx=(20,6))
        # Let column 4 expand so label has space
        frame.grid_columnconfigure(4, weight=1)

    def build_freq_group(self):
        frame = tk.LabelFrame(self.main, text="Frequency Range (for Frequency Response)", bg=FRAME_BG, fg=TEXT_FG, font=FONT_UI, padx=8, pady=8)
        frame.pack(fill="x", pady=(0,8))
        lbl_font = ("Segoe UI", 11)
        tk.Label(frame, text="Min F (Hz):", bg=FRAME_BG, fg=TEXT_FG, font=lbl_font).grid(row=0, column=0, sticky="e", padx=(4,6))
        self.minf_entry = tk.Entry(frame, width=14, bg=INPUT_YELLOW, fg=INPUT_FG, font=("Helvetica",12), insertbackground=INPUT_FG); self.minf_entry.insert(0, "20"); self.minf_entry.grid(row=0, column=1, sticky="w", padx=(0,12))
        tk.Label(frame, text="Max F (Hz):", bg=FRAME_BG, fg=TEXT_FG, font=lbl_font).grid(row=0, column=2, sticky="e", padx=(4,6))
        self.maxf_entry = tk.Entry(frame, width=14, bg=INPUT_YELLOW, fg=INPUT_FG, font=("Helvetica",12), insertbackground=INPUT_FG); self.maxf_entry.insert(0, "20000"); self.maxf_entry.grid(row=0, column=3, sticky="w", padx=(0,12))

    def build_action_group(self):
        frame = tk.Frame(self.main, bg=FRAME_BG)
        frame.pack(fill="x", pady=(0,8))
        self.sweepR_btn = tk.Button(frame, text="Sweep R", bg=BTN_BG, fg=TEXT_FG, width=12, command=self.do_sweep_R); self.sweepR_btn.grid(row=0, column=0, padx=6)
        self.sweepC_btn = tk.Button(frame, text="Sweep C", bg=BTN_BG, fg=TEXT_FG, width=12, command=self.do_sweep_C); self.sweepC_btn.grid(row=0, column=1, padx=6)
        self.sweepL_btn = tk.Button(frame, text="Sweep L", bg=BTN_BG, fg=TEXT_FG, width=12, command=self.do_sweep_L); self.sweepL_btn.grid(row=0, column=2, padx=6)
        self.calc_btn = tk.Button(frame, text="Calculate Fc & Plot", bg="#4CAF50", fg="white", width=18, command=self.calculate_and_plot); self.calc_btn.grid(row=0, column=3, padx=12)
        self.flip_var = tk.BooleanVar(value=False)
        self.flip_cb = tk.Checkbutton(frame, text="Flip X/Y Axes", variable=self.flip_var, bg=FRAME_BG, fg=TEXT_FG, selectcolor=FRAME_BG); self.flip_cb.grid(row=0, column=4, padx=12)
        tk.Label(frame, text="Plot Mode:", bg=FRAME_BG, fg=TEXT_FG).grid(row=0, column=5, padx=(6,4))
        self.plot_mode = tk.StringVar(value="Sweep")
        pmenu = ttk.Combobox(frame, textvariable=self.plot_mode, values=["Sweep","Frequency Response"], width=16, state="readonly"); pmenu.grid(row=0, column=6, sticky="w", padx=(0,12))
        self.show_formula_shown = False
        self.show_formula_btn = tk.Button(frame, text="Show Formulas", bg=BTN_BG, fg=TEXT_FG, width=14, command=self.toggle_formula_area); self.show_formula_btn.grid(row=0, column=7, padx=6)
        self.export_btn = tk.Button(frame, text="Export to Excel", bg=BTN_LIGHT, fg="black", width=16, command=self.export_excel, state="disabled"); self.export_btn.grid(row=0, column=8, padx=6)

    def build_plot_tabs(self):
        tabframe = tk.Frame(self.main, bg=FRAME_BG)
        tabframe.pack(fill="both", expand=True, pady=(0,8))
        self.tabs = ttk.Notebook(tabframe)
        self.tabs.pack(fill="both", expand=True)

        # Magnitude
        mag_frame = tk.Frame(self.tabs, bg=FRAME_BG); self.tabs.add(mag_frame, text="Magnitude (dB)")
        self.fig_mag = Figure(figsize=(7,2), dpi=100); self.ax_mag = self.fig_mag.add_subplot(111); self.ax_mag.grid(True, color="#3b5f7f", linestyle="--", alpha=0.5)
        self.canvas_mag = FigureCanvasTkAgg(self.fig_mag, master=mag_frame); self.canvas_mag.get_tk_widget().pack(fill="both", expand=True)

        # Phase
        phase_frame = tk.Frame(self.tabs, bg=FRAME_BG); self.tabs.add(phase_frame, text="Phase (°)")
        self.fig_phase = Figure(figsize=(7,2), dpi=100); self.ax_phase = self.fig_phase.add_subplot(111); self.ax_phase.grid(True, color="#3b5f7f", linestyle="--", alpha=0.5)
        self.canvas_phase = FigureCanvasTkAgg(self.fig_phase, master=phase_frame); self.canvas_phase.get_tk_widget().pack(fill="both", expand=True)

        # Bode
        bode_frame = tk.Frame(self.tabs, bg=FRAME_BG); self.tabs.add(bode_frame, text="Bode (Mag+Phase)")
        self.fig_bode = Figure(figsize=(7,3.5), dpi=100); self.ax_bode_mag = self.fig_bode.add_subplot(211); self.ax_bode_phase = self.fig_bode.add_subplot(212, sharex=self.ax_bode_mag)
        self.ax_bode_mag.grid(True, color="#3b5f7f", linestyle="--", alpha=0.5); self.ax_bode_phase.grid(True, color="#3b5f7f", linestyle="--", alpha=0.5)
        self.canvas_bode = FigureCanvasTkAgg(self.fig_bode, master=bode_frame); self.canvas_bode.get_tk_widget().pack(fill="both", expand=True)

        # 3D
        view3d_frame = tk.Frame(self.tabs, bg=FRAME_BG); self.tabs.add(view3d_frame, text="3D View")
        self.fig_3d = Figure(figsize=(7,3.5), dpi=100); self.ax_3d = self.fig_3d.add_subplot(111, projection='3d')
        self.canvas_3d = FigureCanvasTkAgg(self.fig_3d, master=view3d_frame); self.canvas_3d.get_tk_widget().pack(fill="both", expand=True)

        # connect cursors
        self.canvas_mag.mpl_connect("motion_notify_event", self.on_mouse_move_mag)
        self.canvas_phase.mpl_connect("motion_notify_event", self.on_mouse_move_phase)
        self.canvas_bode.mpl_connect("motion_notify_event", self.on_mouse_move_bode)
        self.canvas_3d.mpl_connect("motion_notify_event", self.on_mouse_move_3d)
        self.tabs.bind("<<NotebookTabChanged>>", self.on_tab_change)

    def build_table_and_info_with_pane(self):
        """
        Build bottom area as a PanedWindow (resizable sash) containing the table
        and an info/formula panel (collapsible).
        """
        bottom_paned = tk.PanedWindow(self.main, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, bg=FRAME_BG)
        bottom_paned.pack(fill="both", expand=True)

        # Table pane
        table_frame = tk.Frame(bottom_paned, bg=FRAME_BG)
        bottom_paned.add(table_frame, minsize=300)

        self.table_scroll = tk.Scrollbar(table_frame)
        self.table_scroll.pack(side="right", fill="y")

        cols = ("No","R_kΩ","C_μF","L_mH","fc_Hz","Gain")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings", yscrollcommand=self.table_scroll.set, height=12)
        headings = ["#","R (kΩ)","C (μF)","L (mH)","fc / f0 (Hz)","Gain (dB (V))"]
        for c,h in zip(cols,headings):
            self.table.heading(c, text=h); self.table.column(c, width=120, anchor="center")
        self.table.pack(side="left", fill="both", expand=True)
        self.table_scroll.config(command=self.table.yview)

        # Info pane (resizable), holds result label (top-right already) and formula frame collapsible
        info_frame = tk.Frame(bottom_paned, bg=FRAME_BG, width=360)
        bottom_paned.add(info_frame, minsize=200)

        # The formula_frame will be placed inside info_frame when shown.
        self.formula_frame = tk.LabelFrame(info_frame, text="Formulas", bg=FRAME_BG, fg=TEXT_FG, font=FONT_UI, padx=6, pady=6)
        self.formula_text_widget = tk.Text(self.formula_frame, bg=FRAME_BG, fg="#AEDFF6", font=("Courier New",10), width=50, height=16)
        self.formula_text_widget.pack(fill="both", expand=True)
        self.formula_text_widget.config(state="disabled")
        # start hidden (not packed into info_frame)

        # store references
        self.bottom_paned = bottom_paned
        self.info_frame = info_frame
        self.table_frame = table_frame

    def build_3d_controls(self):
        ctrl = tk.Frame(self.main, bg=FRAME_BG)
        ctrl.pack(fill="x", pady=(6,8))
        tk.Label(ctrl, text="3D Sweep Variable:", bg=FRAME_BG, fg=TEXT_FG).grid(row=0, column=0, sticky="e", padx=(4,6))
        self.sweep_var_3d = tk.StringVar(value="R")
        sweep_var_menu = ttk.Combobox(ctrl, textvariable=self.sweep_var_3d, values=["R","C","L"], width=6, state="readonly"); sweep_var_menu.grid(row=0, column=1, sticky="w", padx=(0,12))
        tk.Label(ctrl, text="3D Plot Type:", bg=FRAME_BG, fg=TEXT_FG).grid(row=0, column=2, sticky="e", padx=(4,6))
        self.plot3d_type = tk.StringVar(value="Surface")
        plot3d_menu = ttk.Combobox(ctrl, textvariable=self.plot3d_type, values=["Surface","Waterfall","Phase Surface"], width=14, state="readonly"); plot3d_menu.grid(row=0, column=3, sticky="w", padx=(0,12))
        tk.Label(ctrl, text="Freq Points:", bg=FRAME_BG, fg=TEXT_FG).grid(row=0, column=4, sticky="e", padx=(4,6))
        self.freq_points_3d = tk.Entry(ctrl, width=6, bg=INPUT_YELLOW, fg=INPUT_FG); self.freq_points_3d.insert(0, "80"); self.freq_points_3d.grid(row=0, column=5, sticky="w", padx=(0,12))
        tk.Label(ctrl, text="Sweep Points:", bg=FRAME_BG, fg=TEXT_FG).grid(row=0, column=6, sticky="e", padx=(4,6))
        self.sweep_points_3d = tk.Entry(ctrl, width=6, bg=INPUT_YELLOW, fg=INPUT_FG); self.sweep_points_3d.insert(0, "40"); self.sweep_points_3d.grid(row=0, column=7, sticky="w", padx=(0,12))
        self.gen3d_btn = tk.Button(ctrl, text="Generate 3D", bg=BTN_BG, fg=TEXT_FG, width=14, command=self.generate_3d_plot); self.gen3d_btn.grid(row=0, column=8, padx=6)

    # ---------- Input readers ----------
    def get_R(self):
        raw = self.r_entry.get().strip()
        if raw == "": raise ValueError("R empty")
        v = float(raw)
        unit = self.r_unit.get()
        if unit == "Ω": return v
        if unit == "kΩ": return v * 1e3
        if unit == "MΩ": return v * 1e6
        return v

    def get_C(self):
        raw = self.c_entry.get().strip()
        if raw == "": raise ValueError("C empty")
        v = float(raw)
        unit = self.c_unit.get()
        if unit == "pF": return v * 1e-12
        if unit == "nF": return v * 1e-9
        if unit == "μF": return v * 1e-6
        return v

    def get_L(self):
        raw = self.l_entry.get().strip()
        if raw == "": raise ValueError("L empty")
        v = float(raw)
        unit = self.l_unit.get()
        if unit == "nH": return v * 1e-9
        if unit == "μH": return v * 1e-6
        if unit == "mH": return v * 1e-3
        if unit == "H": return v
        return v

    def get_min_max(self):
        try:
            mn = float(self.minf_entry.get()); mx = float(self.maxf_entry.get())
            if mn <=0 or mx <=0 or mx <= mn: raise ValueError("Invalid min/max")
            return mn, mx
        except Exception as e:
            raise ValueError("Invalid frequency range: " + str(e))

    # ---------- Sweep compute ----------
    def sweep_values(self, sweep_var):
        if sweep_var == 'R': return np.logspace(0, 8, 500)
        if sweep_var == 'C': values = np.logspace(-12, -6, 500)  # in F
        return np.logspace(-9, 0, 500)

    def compute_sweep(self, mode, sweep_var, R_fixed, C_fixed, L_fixed):
        var_vals = self.sweep_values(sweep_var)
        fc_arr = np.full_like(var_vals, np.nan, dtype=float)
        bw_arr = np.full_like(var_vals, np.nan, dtype=float)
        q_arr = np.full_like(var_vals, np.nan, dtype=float)
        gain_arr = np.full_like(var_vals, np.nan, dtype=float)
        phase_arr = np.full_like(var_vals, np.nan, dtype=float)
        for i, v in enumerate(var_vals):
            if sweep_var == 'R': R = v; C = C_fixed; L = L_fixed
            elif sweep_var == 'C': C = v; R = R_fixed; L = L_fixed
            else: L = v; R = R_fixed; C = C_fixed

            if mode == "RC":
                if R>0 and C>0:
                    fc = fc_RC(R,C); fc_arr[i]=fc
                    if self.filter_type_var.get()=="Low-Pass": H,phi = H_RC_lowpass(R,C,np.array([fc]))
                    else: H,phi = H_RC_highpass(R,C,np.array([fc]))
                    gain_arr[i]=H[0]; phase_arr[i]=phi[0]
            elif mode == "RL":
                if R>0 and L>0:
                    fc = fc_RL(R,L); fc_arr[i]=fc
                    if self.filter_type_var.get()=="Low-Pass": H,phi = H_RL_lowpass(R,L,np.array([fc]))
                    else: H,phi = H_RL_highpass(R,L,np.array([fc]))
                    gain_arr[i]=H[0]; phase_arr[i]=phi[0]
            else:
                if L>0 and C>0:
                    f0=f0_RLC(L,C); fc_arr[i]=f0
                    bw_arr[i]=BW_RLC(R,L); q_arr[i]=Q_RLC(R,L,C)
                    gain_arr[i]=H[0]; phase_arr[i]=phi[0]
        return var_vals, fc_arr, bw_arr, q_arr, gain_arr, phase_arr

    # ---------- Display sweep ----------
    def display_sweep(self, var_name, var_vals, fc_vals, bw_vals, q_vals, gain_vals, phase_vals):
        self.last_var_vals = var_vals; self.last_fc = fc_vals; self.last_bw = bw_vals; self.last_q = q_vals
        self.last_gain = gain_vals; self.last_phase = phase_vals; self.last_var_name = var_name
        self.export_ready = True; self.export_btn.config(state="normal")
        self.formula_text_str = self.build_formula_text()
        if self.show_formula_shown: self.formula_text_widget_update()
        self.clear_plots()

        if self.plot_mode.get() == "Sweep":
            if not self.flip_var.get():
                x = var_vals; y = fc_vals; xlabel = var_name; ylabel = "fc (Hz)"
            else:
                x = fc_vals; y = var_vals; xlabel = "fc (Hz)"; ylabel = var_name
            mask = ~np.isnan(x) & ~np.isnan(y)
            if np.any(mask):
                if 'R' in var_name or 'L' in var_name:
                    self.ax_mag.semilogx(x[mask], y[mask], '-o', markersize=3); self.ax_phase.semilogx(x[mask], y[mask], '-o', markersize=3)
                else:
                    self.ax_mag.loglog(x[mask], y[mask], '-o', markersize=3); self.ax_phase.loglog(x[mask], y[mask], '-o', markersize=3)
                self.ax_bode_mag.plot(x[mask], y[mask], '-o', markersize=3); self.ax_bode_phase.plot(x[mask], y[mask], '-o', markersize=3)
            self.ax_mag.set_xlabel(xlabel); self.ax_mag.set_ylabel(ylabel)
            self.ax_phase.set_xlabel(xlabel); self.ax_phase.set_ylabel(ylabel)
        else:
            try:
                R = self.get_R(); C = self.get_C(); L = self.get_L()
                self.plot_frequency_response(R,C,L, show_on_canvas=False)
            except Exception:
                pass

        # update table
        self.table.delete(*self.table.get_children())
        for i in range(len(var_vals)):
            if 'R' in var_name:
                r_display = fmt_number(var_vals[i]/1000.0); c_display = fmt_number(self.get_C()*1e6); l_display = fmt_number(self.get_L()*1e3)
            elif 'C' in var_name:
                r_display = fmt_number(self.get_R()/1000.0); c_display = fmt_number(var_vals[i]*1e6); l_display = fmt_number(self.get_L()*1e3)
            else:
                r_display = fmt_number(self.get_R()/1000.0); c_display = fmt_number(self.get_C()*1e6); l_display = fmt_number(var_vals[i]*1e3)

            fc_show = "-" if np.isnan(fc_vals[i]) else fmt_number(fc_vals[i])
            g_lin = gain_vals[i] if not np.isnan(gain_vals[i]) else np.nan
            g_db = 20*np.log10(g_lin) if g_lin>0 else ( -150.0 if not np.isnan(g_lin) else np.nan )
            gain_show = f"{g_db:.3f} dB ({g_lin:.3f} V)" if not np.isnan(g_lin) else "-"
            bw_show = "-" if np.isnan(bw_vals[i]) else fmt_number(bw_vals[i])
            q_show = "-" if np.isnan(q_vals[i]) else fmt_number(q_vals[i])
            if self.mode_var.get() != "RLC" or self.filter_type_var.get():
                bw_show = "-"; q_show = "-"
            self.table.insert("", "end", values=(i+1, r_display, c_display, l_display, fc_show, gain_show, bw_show, q_show))

        self.canvas_mag.draw(); self.canvas_phase.draw(); self.canvas_bode.draw()

    # ---------- Frequency response ----------
    def plot_frequency_response(self, R, C, L, show_on_canvas=True):
        mn, mx = self.get_min_max()
        f = np.logspace(np.log10(mn), np.log10(mx), 1200)
        H = np.zeros_like(f); phase = np.zeros_like(f)
        mode = self.mode_var.get(); filt = self.filter_type_var.get()

        if mode == "RC":
            if filt == "Low-Pass": H, phase = H_RC_lowpass(R,C,f)
            elif filt == "High-Pass": H, phase = H_RC_highpass(R,C,f)
            else:
                Hlp,phlp = H_RC_lowpass(R,C,f); Hhp,phhp = H_RC_highpass(R,C,f); H = np.abs(Hlp-Hhp); phase = phlp-phhp
        elif mode == "RL":
            if filt == "Low-Pass": H, phase = H_RL_lowpass(R,L,f)
            elif filt == "High-Pass": H, phase = H_RL_highpass(R,L,f)
            else: H, phase = H_RL_lowpass(R,L,f)

        H_db = 20*np.log10(np.maximum(H, 1e-12))
        phase_wrapped = ((phase + 180) % 360) - 180

        self.ax_mag.clear(); self.ax_phase.clear(); self.ax_bode_mag.clear(); self.ax_bode_phase.clear()
        for ax in (self.ax_mag, self.ax_phase, self.ax_bode_mag, self.ax_bode_phase):
            ax.grid(True, color="#3b5f7f", linestyle="--", alpha=0.5)

        if not self.flip_var.get():
            self.ax_mag.semilogx(f, H_db); self.ax_mag.set_xlabel("Frequency (Hz)"); self.ax_mag.set_ylabel("Gain (dB)")
            self.ax_phase.semilogx(f, phase_wrapped); self.ax_phase.set_xlabel("Frequency (Hz)"); self.ax_phase.set_ylabel("Phase (°)")
            self.ax_bode_mag.semilogx(f, H_db); self.ax_bode_mag.set_ylabel("Gain (dB)")
            self.ax_bode_phase.semilogx(f, phase_wrapped); self.ax_bode_phase.set_xlabel("Frequency (Hz)"); self.ax_bode_phase.set_ylabel("Phase (°)")
        else:
            self.ax_mag.plot(H_db, f); self.ax_mag.set_xlabel("Gain (dB)"); self.ax_mag.set_ylabel("Frequency (Hz)")
            self.ax_phase.plot(phase_wrapped, f); self.ax_phase.set_xlabel("Phase (°)"); self.ax_phase.set_ylabel("Frequency (Hz)")
            self.ax_bode_mag.plot(H_db, f); self.ax_bode_mag.set_xlabel("Gain (dB)"); self.ax_bode_mag.set_ylabel("Frequency (Hz)")
            self.ax_bode_phase.plot(phase_wrapped, f); self.ax_bode_phase.set_xlabel("Phase (°)"); self.ax_bode_phase.set_ylabel("Frequency (Hz)")

        self.current_freqs = f; self.current_H_db = H_db; self.current_H_lin = H; self.current_phase = phase_wrapped
        if show_on_canvas:
            self.canvas_mag.draw(); self.canvas_phase.draw(); self.canvas_bode.draw()
        return f, H_db, phase_wrapped

    # ---------- Calculate ----------
    def calculate_and_plot(self):
        try:
            R = self.get_R(); C = self.get_C(); L = self.get_L()
        except Exception as e:
            messagebox.showerror("Input Error", str(e)); return

        mode = self.mode_var.get()
        if mode == "RC": fc = fc_RC(R,C)
        elif mode == "RL": fc = fc_RL(R,L)
        else: fc = f0_RLC(L,C)

        farr, Hdb, phase = self.plot_frequency_response(R,C,L, show_on_canvas=False)
        if np.isnan(fc):
            gain_at_fc_db = None; gain_at_fc_lin = None; phase_at_fc = None
        else:
            idx = np.abs(farr - fc).argmin()
            gain_at_fc_db = Hdb[idx]; gain_at_fc_lin = self.current_H_lin[idx]; phase_at_fc = phase[idx]

        fc_text = human_freq(fc) if not np.isnan(fc) else "N/A"
        gain_text = f"{gain_at_fc_db:.3f} dB ({gain_at_fc_lin:.3f} V)" if gain_at_fc_db is not None else "N/A"
        phase_text = f"{phase_at_fc:.3f} °" if phase_at_fc is not None else "N/A"
        # Update the TOP-RIGHT result label (next to the inputs)
        self.top_result_label.config(text=f"Cutoff / Resonant:\n{fc_text}\n\nGain@fc:\n{gain_text}\n\nPhase@fc:\n{phase_text}")

        # update formula text string and visible widget if shown
        self.formula_text_str = self.build_formula_text()
        if self.show_formula_shown:
            self.formula_text_widget_update()

        # draw canvases
        self.canvas_mag.draw(); self.canvas_phase.draw(); self.canvas_bode.draw()

    # ---------- Formula text ----------
    def build_formula_text(self):
        try:
            R = self.get_R(); C = self.get_C(); L = self.get_L()
        except Exception:
            R=C=L=0.0
        mode = self.mode_var.get(); ftype = self.filter_type_var.get()
        lines = []
        if mode == "RC":
            lines.append("RC - cutoff")
            lines.append("  f_c = 1 / (2π R C)")
            lines.append(f"  Plugged: R = {fmt_number(R)} Ω, C = {fmt_number(C)} F")
            if R>0 and C>0: lines.append(f"  f_c ≈ {human_freq(fc_RC(R,C))}")
        elif mode == "RL":
            lines.append("RL - cutoff")
            lines.append("  f_c = R / (2π L)")
            lines.append(f"  Plugged: R = {fmt_number(R)} Ω, L = {fmt_number(L)} H")
            if R>0 and L>0: lines.append(f"  f_c ≈ {human_freq(fc_RL(R,L))}")
        else:
            lines.append("RLC resonance & damping")
            lines.append("  f0 = 1 / (2π √(L C))")
            lines.append("  BW = R / (2π L)")
            lines.append("  Q = f0 / BW")
            lines.append(f"  Plugged: R = {fmt_number(R)} Ω, L = {fmt_number(L)} H, C = {fmt_number(C)} F")
            if L>0 and C>0: lines.append(f"  f0 ≈ {human_freq(f0_RLC(L,C))}")
            if L>0: lines.append(f"  BW ≈ {human_freq(BW_RLC(R,L))}")
            if L>0 and C>0: lines.append(f"  Q ≈ {fmt_number(Q_RLC(R,L,C))}")
        lines.append("")
        lines.append("Gain magnitude (symbolic):")
        if ftype == "Low-Pass": lines.append("  |H(f)| = 1 / √(1 + (f/f_c)^2)")
        elif ftype == "High-Pass": lines.append("  |H(f)| = (f/f_c) / √(1 + (f/f_c)^2)")
        else: lines.append("  |H(f)| ≈ |1 - (f/f0)^2| / √[(1 - (f/f0)^2)^2 + (f/f0)^2 / Q^2]")
        # also include formula for gain in volts (Vin=1)
        lines.append("")
        lines.append("Gain (voltage) at frequency f: Vout = |H(f)| (assumes Vin = 1 V)")
        return "\n".join(lines)

    def toggle_formula_area(self):
        if not self.show_formula_shown:
            self.formula_text_widget_update()
            # pack the formula_frame into the info_frame (which is the second pane in the bottom paned window)
            self.formula_frame.pack(fill="both", expand=True)
            self.show_formula_btn.config(text="Hide Formulas")
            self.show_formula_shown = True
        else:
            self.formula_frame.pack_forget()
            self.show_formula_btn.config(text="Show Formulas")
            self.show_formula_shown = False

    def formula_text_widget_update(self):
        self.formula_text_widget.config(state="normal")
        self.formula_text_widget.delete("1.0","end")
        self.formula_text_widget.insert("1.0", self.formula_text_str or self.build_formula_text())
        self.formula_text_widget.config(state="disabled")

    # ---------- Sweep handlers ----------
    def do_sweep_R(self):
        try:
            R_fixed = self.get_R(); C_fixed = self.get_C(); L_fixed = self.get_L()
        except Exception as e:
            messagebox.showerror("Input Error", str(e)); return
        mode = self.mode_var.get()
        var_vals, fc, bw, q, gain, phase = self.compute_sweep(mode, 'R', R_fixed, C_fixed, L_fixed)
        self.display_sweep('R (Ω)', var_vals, fc, bw, q, gain, phase)

    def do_sweep_C(self):
        try:
            R_fixed = self.get_R(); C_fixed = self.get_C(); L_fixed = self.get_L()
        except Exception as e:
            messagebox.showerror("Input Error", str(e)); return
        mode = self.mode_var.get()
        var_vals, fc, bw, q, gain, phase = self.compute_sweep(mode, 'C', R_fixed, C_fixed, L_fixed)
        self.display_sweep('C (F)', var_vals, fc, bw, q, gain, phase)

    def do_sweep_L(self):
        try:
            R_fixed = self.get_R(); C_fixed = self.get_C(); L_fixed = self.get_L()
        except Exception as e:
            messagebox.showerror("Input Error", str(e)); return
        mode = self.mode_var.get()
        var_vals, fc, bw, q, gain, phase = self.compute_sweep(mode, 'L', R_fixed, C_fixed, L_fixed)
        self.display_sweep('L (H)', var_vals, fc, bw, q, gain, phase)

    # ---------- Cursor & tooltip ----------
    def on_tab_change(self, event):
        self.canvas_mag.draw(); self.canvas_phase.draw(); self.canvas_bode.draw(); self.canvas_3d.draw()

    def clear_plots(self):
        # ensure axes exist
        self.ax_mag = self.fig_mag.axes[0] if self.fig_mag.axes else self.fig_mag.add_subplot(111)
        self.ax_phase = self.fig_phase.axes[0] if self.fig_phase.axes else self.fig_phase.add_subplot(111)
        if len(self.fig_bode.axes) < 2:
            self.fig_bode.clf(); self.ax_bode_mag = self.fig_bode.add_subplot(211); self.ax_bode_phase = self.fig_bode.add_subplot(212, sharex=self.ax_bode_mag)
        else:
            self.ax_bode_mag = self.fig_bode.axes[0]; self.ax_bode_phase = self.fig_bode.axes[1]
        self.ax_3d = self.fig_3d.axes[0] if self.fig_3d.axes else self.fig_3d.add_subplot(111, projection='3d')

        for ax in (self.ax_mag, self.ax_phase, self.ax_bode_mag, self.ax_bode_phase):
            ax.clear(); ax.grid(True, color="#3b5f7f", linestyle="--", alpha=0.5)
        self.ax_3d.clear()

        self.canvas_mag.draw(); self.canvas_phase.draw(); self.canvas_bode.draw(); self.canvas_3d.draw()
        self.current_freqs = np.array([]); self.current_H_db = np.array([]); self.current_H_lin = np.array([]); self.current_phase = np.array([])

    def _draw_cursor_and_tooltip(self, ax, event):
        if event.xdata is None: return
        x = event.xdata
        # remove old cursor/text
        for ln in list(ax.lines):
            if getattr(ln, "_is_cursor", False): ln.remove()
        for t in list(ax.texts):
            if getattr(t, "_is_tooltip", False): t.remove()
        # vertical line
        ln = ax.axvline(x, color="cyan", linewidth=1, linestyle="--"); ln._is_cursor = True
        # interpolate values
        if hasattr(self, "current_freqs") and self.current_freqs.size>0:
            f = self.current_freqs; xcl = np.clip(x, f[0], f[-1])
            Hdb = np.interp(xcl, f, self.current_H_db); Hlin = np.interp(xcl, f, self.current_H_lin); phase = np.interp(xcl, f, self.current_phase)
            txt = f"f = {xcl:.3f} Hz\nGain = {Hdb:.3f} dB ({Hlin:.3f} V)\nPhase = {phase:.3f}°"
        else:
            txt = f"x = {x:.3f}"
        tooltip = ax.text(0.02, 0.95, txt, transform=ax.transAxes, fontsize=9, verticalalignment='top', bbox=dict(boxstyle="round", fc="#033a56", ec="cyan", alpha=0.9))
        tooltip._is_tooltip = True

    def on_mouse_move_mag(self, event):
        self._draw_cursor_and_tooltip(self.ax_mag, event); self.canvas_mag.draw_idle()

    def on_mouse_move_phase(self, event):
        self._draw_cursor_and_tooltip(self.ax_phase, event); self.canvas_phase.draw_idle()

    def on_mouse_move_bode(self, event):
        if event.xdata is None: return
        x = event.xdata
        for ax in (self.ax_bode_mag, self.ax_bode_phase):
            for ln in list(ax.lines):
                if getattr(ln, "_is_cursor", False): ln.remove()
            for t in list(ax.texts):
                if getattr(t, "_is_tooltip", False): t.remove()
        for ax in (self.ax_bode_mag, self.ax_bode_phase):
            ln = ax.axvline(x, color="cyan", linewidth=1, linestyle="--"); ln._is_cursor = True
        if hasattr(self, "current_freqs") and self.current_freqs.size>0:
            f = self.current_freqs; xcl = np.clip(x, f[0], f[-1])
            Hdb = np.interp(xcl, f, self.current_H_db); Hlin = np.interp(xcl, f, self.current_H_lin); phase = np.interp(xcl, f, self.current_phase)
            txt = f"f = {xcl:.3f} Hz\nGain = {Hdb:.3f} dB ({Hlin:.3f} V)\nPhase = {phase:.3f}°"
        else:
            txt = f"x = {x:.3f}"
        tooltip = self.ax_bode_mag.text(0.02, 0.95, txt, transform=self.ax_bode_mag.transAxes, verticalalignment='top', bbox=dict(boxstyle="round", fc="#033a56", ec="cyan", alpha=0.9))
        tooltip._is_tooltip = True
        self.canvas_bode.draw_idle()

    def on_mouse_move_3d(self, event):
        ax = self.ax_3d
        for t in list(ax.texts):
            if getattr(t, "_is_tooltip", False): t.remove()
        txt = f"x={event.x:.0f}, y={event.y:.0f}"
        tooltip = ax.text2D(0.02, 0.95, txt, transform=ax.transAxes, fontsize=9, bbox=dict(boxstyle="round", fc="#033a56", ec="cyan", alpha=0.9))
        tooltip._is_tooltip = True
        self.canvas_3d.draw_idle()

    # ---------- 3D generation ----------
    def generate_3d_plot(self):
        var = self.sweep_var_3d.get(); ptype = self.plot3d_type.get()
        try:
            freq_pts = int(self.freq_points_3d.get()); sweep_pts = int(self.sweep_points_3d.get()); mn, mx = self.get_min_max()
            R0 = self.get_R(); C0 = self.get_C(); L0 = self.get_L()
        except Exception as e:
            messagebox.showerror("Invalid input", str(e)); return

        f = np.logspace(np.log10(mn), np.log10(mx), freq_pts)
        if var=="R": var_vals = np.logspace(0, 8, sweep_pts)
        elif var=="C": var_vals = np.logspace(-12, -6, sweep_pts)
        else: var_vals = np.logspace(-9, 0, sweep_pts)

        F_mesh, V_mesh = np.meshgrid(f, var_vals)
        H_mesh = np.zeros_like(F_mesh); Phi_mesh = np.zeros_like(F_mesh)
        mode = self.mode_var.get(); filt = self.filter_type_var.get()

        for i in range(V_mesh.shape[0]):
            val = V_mesh[i,0]
            if var == "R": R = val; C = C0; L = L0
            elif var == "C": R = R0; C = val; L = L0
            else: R = R0; C = C0; L = val

            if mode == "RC":
                if filt == "Low-Pass": H_row, Phi_row = H_RC_lowpass(R,C,f)
                elif filt == "High-Pass": H_row, Phi_row = H_RC_highpass(R,C,f)
                else: Hlp,phlp = H_RC_lowpass(R,C,f); Hhp,phhp = H_RC_highpass(R,C,f); H_row = np.abs(Hlp-Hhp); Phi_row = phlp-phhp
            elif mode == "RL":
                if filt == "Low-Pass": H_row, Phi_row = H_RL_lowpass(R,L,f)
                elif filt == "High-Pass": H_row, Phi_row = H_RL_highpass(R,L,f)
                else: H_row, Phi_row = H_RL_lowpass(R,L,f)

            H_mesh[i,:] = 20*np.log10(np.maximum(H_row, 1e-12))
            Phi_mesh[i,:] = ((Phi_row + 180) % 360) - 180

        self.ax_3d.clear(); self.ax_3d.set_facecolor("#053A6A"); self.ax_3d.grid(True, color="#3b5f7f", linestyle="--", alpha=0.5)
        if ptype == "Surface":
            X = np.log10(F_mesh); Y = np.log10(V_mesh); Z = H_mesh
            surf = self.ax_3d.plot_surface(X, Y, Z, cmap='viridis', linewidth=0, antialiased=True)
            self.ax_3d.set_xlabel("log10(Freq Hz)"); self.ax_3d.set_ylabel(f"log10({var})"); self.ax_3d.set_zlabel("Gain (dB)")
            self.fig_3d.colorbar(surf, ax=self.ax_3d, shrink=0.5, aspect=12)
        elif ptype == "Waterfall":
            for i in range(H_mesh.shape[0]):
                Ypos = np.log10(var_vals[i])
                X = f; Z = H_mesh[i,:]
                self.ax_3d.plot(np.log10(X), np.full_like(X, Ypos), Z, color=cm.viridis(i / H_mesh.shape[0]))
            self.ax_3d.set_xlabel("log10(Freq Hz)"); self.ax_3d.set_ylabel(f"log10({var})"); self.ax_3d.set_zlabel("Gain (dB)")
        else:
            X = np.log10(F_mesh); Y = np.log10(V_mesh); Z = Phi_mesh
            surf = self.ax_3d.plot_surface(X, Y, Z, cmap='coolwarm', linewidth=0, antialiased=True)
            self.ax_3d.set_xlabel("log10(Freq Hz)"); self.ax_3d.set_ylabel(f"log10({var})"); self.ax_3d.set_zlabel("Phase (°)")
            self.fig_3d.colorbar(surf, ax=self.ax_3d, shrink=0.5, aspect=12)

        self.canvas_3d.draw()

    # ---------- Export ----------
    def export_excel(self):
        if not self.export_ready:
            messagebox.showwarning("No data", "Run a sweep first to enable export."); return
        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default = f"filter_export_{now}.xlsx"
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default, filetypes=[("Excel","*.xlsx")])
        if not file: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Filter Data"
        row = 1
        settings = {
            "Mode": self.mode_var.get(),
            "Filter Type": self.filter_type_var.get(),
            "R": f"{self.r_entry.get()} {self.r_unit.get()}",
            "C": f"{self.c_entry.get()} {self.c_unit.get()}",
            "L": f"{self.l_entry.get()} {self.l_unit.get()}",
            "Min F": self.minf_entry.get(),
            "Max F": self.maxf_entry.get()
        }
        ws.cell(row=row, column=1, value="Settings"); row+=1
        for k,v in settings.items():
            ws.cell(row=row, column=1, value=k); ws.cell(row=row, column=2, value=v); row+=1
        row += 1; ws.cell(row=row, column=1, value="Formula Reference"); row+=1
        for line in (self.formula_text_str or self.build_formula_text()).splitlines():
            ws.cell(row=row, column=1, value=line); row+=1
        row += 1
        headers = [self.last_var_name, "fc (Hz)", "Gain (dB)", "Gain (V)"]
        for c,h in enumerate(headers, start=1): ws.cell(row=row, column=c, value=h)
        row += 1
        for i in range(len(self.last_var_vals)):
            ws.cell(row=row, column=1, value=float(fmt_number(self.last_var_vals[i])))
            fcval = None if np.isnan(self.last_fc[i]) else float(fmt_number(self.last_fc[i])); ws.cell(row=row, column=2, value=fcval)
            g_lin = None if np.isnan(self.last_gain[i]) else float(fmt_number(self.last_gain[i])); g_db = None if g_lin is None else float(fmt_number(20*np.log10(g_lin)))
            ws.cell(row=row, column=3, value=g_db); ws.cell(row=row, column=4, value=g_lin)
            ws.cell(row=row, column=5, value=None if np.isnan(self.last_bw[i]) else float(fmt_number(self.last_bw[i])))
            ws.cell(row=row, column=6, value=None if np.isnan(self.last_q[i]) else float(fmt_number(self.last_q[i])))
            row += 1

        def save_fig_to_sheet(fig, sheetname):
            buf = BytesIO(); fig.savefig(buf, format='png', bbox_inches='tight'); buf.seek(0); img = XLImage(buf); ws2 = wb.create_sheet(title=sheetname); ws2.add_image(img, "A1")

        save_fig_to_sheet(self.fig_mag, "Magnitude"); save_fig_to_sheet(self.fig_phase, "Phase"); save_fig_to_sheet(self.fig_bode, "Bode"); save_fig_to_sheet(self.fig_3d, "3D View")
        wb.save(file); messagebox.showinfo("Export Complete", f"Export saved to:\n{file}")

# ---------- Run ----------
if __name__ == "__main__":
    root = tk.Tk()
    app = FilterTool3DApp(root)
    root.mainloop()
